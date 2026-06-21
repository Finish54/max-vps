import io

from aiogram.types import BufferedInputFile
from openpyxl.styles import Alignment, Font
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


async def get_excel_file(
    columns: list[str],
    field: list[list[str]],
    name_file: str
) -> BufferedInputFile:
    """
    Функция возвращает Excel файл в буфере
    :param columns: Список названий для колонок
    :param field: Список данных по строчкам
    :param name_file: Название файла без расширения
    :return: file: BufferedInputFile
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'BotMetrics'
    worksheet.append(columns)
    for data in field:
        worksheet.append(data)
    await document_setup(23, len(columns), worksheet)
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    file_content = buffer.getvalue()
    file = BufferedInputFile(file_content, f'{name_file}.xlsx')
    return file


async def document_setup(
    width: int,
    length: int,
    worksheet: Worksheet
) -> None:
    # Настройка стиля заголовок и выравнивание по центру
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

        # Настройка фиксированный ширины столбцов с заголовками
    for column in worksheet.columns:
        col_letter = column[0].column_letter
        worksheet.column_dimensions[col_letter].width = width

        # Настройка выравнивая значений в ячейках
    for row in worksheet.iter_rows(
        min_row=2,
        max_row=worksheet.max_row,
        min_col=1,
        max_col=length
    ):
        for cell in row:
            cell.alignment = Alignment(
                horizontal='center',
                wrapText=True,
                vertical='center'
            )
